"""
Phase 1 – Data Ingestion
Reads an .xlsx exam file and appends/updates the persistent CSV store.
"""

import os
import re
import csv
import warnings
import openpyxl

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    SCORE_SHEET_NAME, CLASS_WISE_SHEET_NAME, EXAM_NAME_CELL,
    REQUIRED_COLUMNS, CSV_COLUMNS, CSV_PATH, DATA_DIR,
)


def _get_column_indices(header_row: list) -> dict:
    """
    Return a mapping {column_name: 0-based index} from the header row.
    When a column name appears more than once, only the FIRST occurrence is used.
    """
    seen = set()
    mapping = {}
    for idx, val in enumerate(header_row):
        if val and val not in seen:
            mapping[val] = idx
            seen.add(val)
    return mapping


def _normalise_exam_name(raw: str) -> str:
    """
    Normalise the raw exam name from Class_Wise!A4 into a clean stored label.

    Rules:
      - Contains "CDF" (case-insensitive) → "CDF - {last number}"
      - Anything else                      → "JEE - {last number}"
      - No number found                    → raise ValueError

    Examples:
      "IIT Foundation - 12"           → "JEE - 12"
      "IIT FOUNDATION_LEVEL - 2 : 12" → "JEE - 12"
      "CDF-Test-3"                    → "CDF - 3"
      "CDF LEVEL 2 - 7"              → "CDF - 7"
    """
    numbers = re.findall(r"\d+", raw)
    if not numbers:
        raise ValueError(
            f"Cannot extract a test number from exam name: '{raw}'. "
            "The exam name in Class_Wise cell A4 must contain a number (e.g. 'IIT Foundation - 12' or 'CDF-Test-3')."
        )
    number = numbers[-1]  # always use the last number found
    prefix = "CDF" if re.search(r"CDF", raw, re.IGNORECASE) else "JEE"
    return f"{prefix} - {number}"


def _parse_exam_name(ws_class_wise) -> str:
    """Read and normalise exam name from the Class_Wise sheet cell A4."""
    value = ws_class_wise[EXAM_NAME_CELL].value
    if not value:
        raise ValueError(
            f"Cell {EXAM_NAME_CELL} in sheet '{CLASS_WISE_SHEET_NAME}' is empty. "
            "Cannot determine exam name."
        )
    raw = str(value).strip()
    return _normalise_exam_name(raw)


def _to_numeric(val):
    """
    Convert a raw Excel cell value to int or float, handling custom formats.

    Cases covered:
      - Plain int/float (normal)         → int or float
      - datetime / date object           → extract the Excel serial via toordinal
                                           then attempt numeric parse of the string
      - String with extra characters     → strip non-numeric chars, then parse
        e.g. "45 marks", "45/80"        → 45
      - Anything unparseable             → returned as-is (flagged during ingestion)
    """
    import datetime as _dt
    import re as _re

    # datetime / date → recover the original Excel serial number.
    # Excel stores numbers as date serials (score 45 = serial 45 = 1900-02-14).
    # Excel wrongly treats 1900 as a leap year (phantom Feb 29 = serial 60).
    # openpyxl skips that day, so for serials >= 61 the date is 1 day behind.
    # Fix: add 1 for any date on or after 1900-03-01.
    if isinstance(val, (_dt.datetime, _dt.date)):
        epoch = _dt.date(1899, 12, 31)
        d = val.date() if isinstance(val, _dt.datetime) else val
        serial = (d - epoch).days
        if d >= _dt.date(1900, 3, 1):
            serial += 1
        val = serial

    # Stringify and strip whitespace
    s = str(val).strip()

    # Direct numeric parse
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        pass

    # "45/80" format — take the numerator (part before the slash)
    if "/" in s:
        numerator = s.split("/")[0].strip()
        try:
            f = float(numerator)
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            pass

    # Strip non-numeric characters (keep digits, dot, minus)
    cleaned = _re.sub(r"[^\d.\-]", "", s)
    if cleaned:
        try:
            f = float(cleaned)
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            pass

    # Fallback: return original string (will be visible in CSV for manual review)
    return s


def _is_absent(record: dict) -> bool:
    """
    Return True if this student row is marked absent.
    A row is absent when ANY score column contains 'AB' (case-insensitive).
    Since AB is all-or-nothing, checking one column is sufficient,
    but we check all three for safety.
    """
    score_cols = ["Physics Score", "Chemistry Score", "Mathematics Score"]
    for col in score_cols:
        val = str(record.get(col, "")).strip().upper()
        if val == "AB":
            return True
    return False


def _total_marks(record: dict) -> float:
    """
    Sum the three subject scores for comparison purposes.
    Returns 0.0 if any score is non-numeric (AB or blank).
    """
    total = 0.0
    for col in ["Physics Score", "Chemistry Score", "Mathematics Score"]:
        val = record.get(col, "")
        try:
            total += float(val)
        except (ValueError, TypeError):
            return 0.0
    return total


def _resolve_conflict(existing: dict, incoming: dict) -> tuple[dict, str]:
    """
    Decide which record to keep when (Username, Exam) already exists in the CSV.

    Returns (record_to_store, reason_string).

    Rules:
      - existing=AB,  incoming=values → keep incoming  ("absent overridden by scores")
      - existing=values, incoming=AB  → keep existing  ("kept existing scores, ignored AB")
      - existing=AB,  incoming=AB     → keep incoming  ("both absent, no change")
      - existing=values, incoming=values → keep whichever has the GREATER total marks;
                                           if equal, keep incoming
    """
    existing_absent = _is_absent(existing)
    incoming_absent = _is_absent(incoming)

    if existing_absent and not incoming_absent:
        return incoming, "absent overridden by scores"
    elif not existing_absent and incoming_absent:
        return existing, "kept existing scores, ignored AB"
    elif existing_absent and incoming_absent:
        return incoming, "both absent"
    else:
        # Both have numeric scores — keep the one with greater total
        existing_total = _total_marks(existing)
        incoming_total = _total_marks(incoming)
        if incoming_total >= existing_total:
            return incoming, f"updated with higher/equal marks ({incoming_total} >= {existing_total})"
        else:
            return existing, f"kept existing higher marks ({existing_total} > {incoming_total})"


def _load_existing_csv(csv_path: str) -> tuple[list[dict], dict]:
    """
    Load existing CSV rows.
    Returns (rows_list, lookup_dict) where lookup_dict key = (username, exam).
    """
    rows = []
    lookup = {}
    if not os.path.exists(csv_path):
        return rows, lookup

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))
            key = (row["Username"].strip(), row["Exam"].strip())
            lookup[key] = len(rows) - 1  # index into rows list
    return rows, lookup


def _write_csv(csv_path: str, rows: list[dict]) -> None:
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def ingest_excel(file_path: str, csv_path: str = CSV_PATH) -> None:
    """
    Main ingestion function.
    1. Read the Excel file.
    2. Extract exam name from Class_Wise!A4.
    3. Extract student rows from Score_Sheet.
    4. Append or update the persistent CSV.
    5. Print a summary.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    print(f"Opening: {file_path}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(file_path, data_only=True)

    # ── Validate sheets ────────────────────────────────────────────────────────
    missing = [s for s in [SCORE_SHEET_NAME, CLASS_WISE_SHEET_NAME] if s not in wb.sheetnames]
    if missing:
        raise ValueError(
            f"Required sheet(s) not found: {missing}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws_score = wb[SCORE_SHEET_NAME]
    ws_class = wb[CLASS_WISE_SHEET_NAME]

    # ── Exam name ──────────────────────────────────────────────────────────────
    exam_name = _parse_exam_name(ws_class)
    print(f"Exam name: {exam_name}")

    # ── Header row & column mapping ────────────────────────────────────────────
    header_row = [cell.value for cell in ws_score[1]]
    col_map = _get_column_indices(header_row)

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing_cols:
        raise ValueError(
            f"Missing required columns in '{SCORE_SHEET_NAME}': {missing_cols}\n"
            f"Found columns: {[h for h in header_row if h]}"
        )

    # ── Extract new rows ───────────────────────────────────────────────────────
    new_rows: list[dict] = []
    for row_idx, row in enumerate(ws_score.iter_rows(min_row=2, values_only=True), start=2):
        username = row[col_map["Username"]]
        if not username:
            continue  # skip blank rows

        record = {
            "Username": str(username).strip(),
            "First Name": str(row[col_map["First Name"]] or "").strip(),
            "Batch": str(row[col_map["Batch"]] or "").strip(),
            "Physics Score": row[col_map["Physics Score"]],
            "Chemistry Score": row[col_map["Chemistry Score"]],
            "Mathematics Score": row[col_map["Mathematics Score"]],
            "Exam": exam_name,
        }

        # Convert scores to int/float, or keep "AB" as-is.
        # Handles custom Excel formats: datetime objects, floats, strings with
        # extra characters (e.g. "45 marks", "45/80"), and date serial numbers.
        for score_col in ["Physics Score", "Chemistry Score", "Mathematics Score"]:
            val = record[score_col]
            if val is None or val == "":
                record[score_col] = ""
            elif str(val).strip().upper() == "AB":
                record[score_col] = "AB"
            else:
                record[score_col] = _to_numeric(val)

        new_rows.append(record)

    if not new_rows:
        print("No data rows found in Score_Sheet. Nothing ingested.")
        return

    # ── Load existing CSV & merge ──────────────────────────────────────────────
    existing_rows, lookup = _load_existing_csv(csv_path)

    added = 0
    updated = 0
    skipped_ab = 0
    batches_seen: set[str] = set()

    for record in new_rows:
        key = (record["Username"], record["Exam"])
        batches_seen.add(record["Batch"])
        if key in lookup:
            resolved, reason = _resolve_conflict(existing_rows[lookup[key]], record)
            existing_rows[lookup[key]] = resolved
            if "kept existing" in reason:
                skipped_ab += 1
            else:
                updated += 1
        else:
            lookup[key] = len(existing_rows)
            existing_rows.append(record)
            added += 1

    _write_csv(csv_path, existing_rows)

    # ── Summary ────────────────────────────────────────────────────────────────
    print()
    print("=" * 50)
    print(f"  Ingestion complete for: {exam_name}")
    print(f"  Rows added        : {added}")
    print(f"  Rows updated      : {updated}")
    print(f"  AB kept (no overwrite): {skipped_ab}")
    print(f"  Batches      : {', '.join(sorted(batches_seen))}")
    print(f"  CSV location : {csv_path}")
    print("=" * 50)
